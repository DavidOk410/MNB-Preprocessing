from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def colorize_binary_columns(file_path, output_path=None):
    """
    For every column in the Excel file that contains only 0s and 1s:
      - Fill cells with value 1 → green
      - Fill cells with value 0 → red
    Uses conditional formatting (rules survive re-opens in Excel).
    """
    output_path = output_path or file_path

    wb = load_workbook(file_path)
    ws = wb.active

    green_fill = PatternFill(start_color="50A070", end_color="50A070", fill_type="solid")
    red_fill   = PatternFill(start_color="BB7575", end_color="BB7575", fill_type="solid")

    # Find header row and detect binary columns (only 0s and 1s, no other values)
    headers = [cell.value for cell in ws[1]]
    max_row = ws.max_row

    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter

        # Collect all non-empty values in this column (skip header)
        values = set()
        for row in range(2, max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                values.add(val)

        # Only apply to columns that contain exclusively 0s and 1s
        if values and values.issubset({0, 1}):
            cell_range = f"{col_letter}2:{col_letter}{max_row}"

            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator="equal", formula=["1"], fill=green_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator="equal", formula=["0"], fill=red_fill))

            print(f"  Colorized column: {header} ({col_letter})")

    wb.save(output_path)
    print(f"\nSaved colorized file to: {output_path}")